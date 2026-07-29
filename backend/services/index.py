import base64
import json
import os
from io import BytesIO

import psycopg2
from openpyxl import load_workbook


def get_conn():
    return psycopg2.connect(os.environ['DATABASE_URL'])


def cors_headers():
    return {
        'Access-Control-Allow-Origin': '*',
        'Access-Control-Allow-Methods': 'GET, POST, PUT, DELETE, OPTIONS',
        'Access-Control-Allow-Headers': 'Content-Type, X-Authorization',
        'Access-Control-Max-Age': '86400'
    }


def response(status: int, body):
    return {
        'statusCode': status,
        'headers': {**cors_headers(), 'Content-Type': 'application/json'},
        'body': json.dumps(body, default=str)
    }


def get_current_user(cur, event):
    headers = event.get('headers') or {}
    token = headers.get('X-Authorization') or headers.get('x-authorization') or ''
    token = token.replace('Bearer ', '')
    if not token:
        return None
    cur.execute(
        "SELECT u.id, u.company_id FROM sessions s JOIN users u ON u.id = s.user_id WHERE s.token = %s AND s.expires_at > NOW()",
        (token,)
    )
    row = cur.fetchone()
    if not row:
        return None
    return {'user_id': row[0], 'company_id': row[1]}


def handler(event: dict, context) -> dict:
    '''CRUD справочника услуг компании FixKey'''
    method = event.get('httpMethod', 'GET')

    if method == 'OPTIONS':
        return {'statusCode': 200, 'headers': cors_headers(), 'body': ''}

    conn = get_conn()
    cur = conn.cursor()

    try:
        user = get_current_user(cur, event)
        if not user:
            return response(401, {'error': 'Не авторизован'})
        company_id = user['company_id']

        params = event.get('queryStringParameters') or {}
        service_id = params.get('id')

        action = params.get('action')

        if method == 'GET':
            cur.execute(
                "SELECT id, name, unit, price, category, subcategory, created_at FROM services WHERE company_id = %s ORDER BY created_at DESC",
                (company_id,)
            )
            rows = cur.fetchall()
            keys = ['id', 'name', 'unit', 'price', 'category', 'subcategory', 'created_at']
            return response(200, {'services': [dict(zip(keys, r)) for r in rows]})

        if method == 'POST' and action == 'import':
            body = json.loads(event.get('body') or '{}')
            file_b64 = body.get('file') or ''
            if not file_b64:
                return response(400, {'error': 'Файл не передан'})

            try:
                file_bytes = base64.b64decode(file_b64)
                wb = load_workbook(filename=BytesIO(file_bytes), read_only=True, data_only=True)
                sheet = wb.active
            except Exception:
                return response(400, {'error': 'Не удалось прочитать файл. Убедитесь, что это .xlsx'})

            imported = 0
            skipped = 0
            rows_iter = sheet.iter_rows(min_row=1, values_only=True)
            header = next(rows_iter, None)

            def norm(v):
                return (str(v).strip().lower() if v is not None else '')

            header_map = {}
            if header:
                for idx, col in enumerate(header):
                    key = norm(col)
                    if key in ('название', 'название работы', 'name', 'услуга'):
                        header_map['name'] = idx
                    elif key in ('ед.изм.', 'ед. изм.', 'единица', 'unit', 'ед'):
                        header_map['unit'] = idx
                    elif key in ('цена', 'price', 'стоимость'):
                        header_map['price'] = idx
                    elif key in ('категория', 'category'):
                        header_map['category'] = idx
                    elif key in ('подкатегория', 'subcategory'):
                        header_map['subcategory'] = idx

            has_header = 'name' in header_map
            if not has_header:
                header_map = {'name': 0, 'unit': 1, 'price': 2, 'category': 3, 'subcategory': 4}
                rows_iter = sheet.iter_rows(min_row=1, values_only=True)

            for row in rows_iter:
                if not row or all(c is None for c in row):
                    continue

                def get(field, default=''):
                    idx = header_map.get(field)
                    if idx is None or idx >= len(row) or row[idx] is None:
                        return default
                    return row[idx]

                name = str(get('name', '')).strip()
                if len(name) < 2:
                    skipped += 1
                    continue

                unit = str(get('unit', 'м²')).strip() or 'м²'
                try:
                    price = float(get('price', 0) or 0)
                except (TypeError, ValueError):
                    price = 0
                category = str(get('category', '')).strip()
                subcategory = str(get('subcategory', '')).strip()

                cur.execute(
                    "INSERT INTO services (company_id, name, unit, price, category, subcategory) VALUES (%s, %s, %s, %s, %s, %s)",
                    (company_id, name, unit, price, category, subcategory)
                )
                imported += 1

            conn.commit()
            return response(200, {'success': True, 'imported': imported, 'skipped': skipped})

        if method == 'POST':
            body = json.loads(event.get('body') or '{}')
            name = (body.get('name') or '').strip()
            unit = (body.get('unit') or 'м²').strip()
            price = body.get('price') or 0
            category = (body.get('category') or '').strip()
            subcategory = (body.get('subcategory') or '').strip()

            if len(name) < 2:
                return response(400, {'error': 'Введите название услуги'})

            cur.execute(
                "INSERT INTO services (company_id, name, unit, price, category, subcategory) VALUES (%s, %s, %s, %s, %s, %s) RETURNING id, created_at",
                (company_id, name, unit, price, category, subcategory)
            )
            new_id, created_at = cur.fetchone()
            conn.commit()

            return response(200, {
                'id': new_id, 'name': name, 'unit': unit, 'price': price,
                'category': category, 'subcategory': subcategory, 'created_at': created_at
            })

        if method == 'PUT':
            if not service_id:
                return response(400, {'error': 'Не указан id услуги'})
            body = json.loads(event.get('body') or '{}')

            cur.execute("SELECT id FROM services WHERE id = %s AND company_id = %s", (service_id, company_id))
            if not cur.fetchone():
                return response(404, {'error': 'Услуга не найдена'})

            fields = []
            values = []
            for key in ['name', 'unit', 'price', 'category', 'subcategory']:
                if key in body:
                    fields.append(f"{key} = %s")
                    values.append(body[key])

            if fields:
                values.append(service_id)
                values.append(company_id)
                cur.execute(f"UPDATE services SET {', '.join(fields)} WHERE id = %s AND company_id = %s", values)
                conn.commit()

            return response(200, {'success': True})

        if method == 'DELETE':
            if not service_id:
                return response(400, {'error': 'Не указан id услуги'})
            cur.execute("DELETE FROM services WHERE id = %s AND company_id = %s", (service_id, company_id))
            conn.commit()
            return response(200, {'success': True})

        return response(405, {'error': 'Метод не поддерживается'})
    finally:
        cur.close()
        conn.close()
# touch